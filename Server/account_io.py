# This module handles xlsx files
import openpyxl
from database_management import user_management
class io_manager:
	def read_file(filename = './accounts.xlsx'):
		u_list = []
		p_list = []
		t_list = []
		try:
			workbook_object = openpyxl.load_workbook(filename)
			# get active sheet
			sheet_object = workbook_object.active
			# Get number of rows
			number_of_rows = sheet_object.max_row
		except:
			print('[ ERROR ] Could not open ' + filename)
			return u_list, ['FNF'] , t_list
		
		try:
			for i in range(2, number_of_rows + 1):
				cell_object = sheet_object.cell(row = i, column = 1)
				u_list.append(cell_object.value)

				cell_object = sheet_object.cell(row = i, column = 2)
				p_list.append(cell_object.value)

				cell_object = sheet_object.cell(row = i, column = 3)
				t_list.append(cell_object.value)
			return u_list, p_list, t_list

		except Exception as error:
			print('[ ERROR ] File could not be read properly!')
			return [], [], []

	def write_file(u_list, p_list, t_list, filename = './accounts.xlsx'):
		try:
			workbook_object = openpyxl.Workbook()
			# get active sheet
			sheet_object = workbook_object.active
		except Exception as error:
			print('[ ERROR ] ' + str(error))

		try:
			cell_object = sheet_object.cell(row = 1, column = 1)
			cell_object.value = 'Team'
			cell_object = sheet_object.cell(row = 1, column = 2)
			cell_object.value = 'Password'
			cell_object = sheet_object.cell(row = 1, column = 3)
			cell_object.value = 'Type'

			for i in range(2, 2 + len(u_list)):
				cell_object = sheet_object.cell(row = i, column = 1)
				cell_object.value = u_list[i-2]

				cell_object = sheet_object.cell(row = i, column = 2)
				cell_object.value = p_list[i-2]

				cell_object = sheet_object.cell(row = i, column = 3)
				cell_object.value = t_list[i-2]
			
			workbook_object.save(filename)
		except Exception as error:
			print('[ ERROR ] File could not be read properly: ' + str(error))

if __name__ == '__main__':
	io_manager.write_file(['team1', 'team2'], ['abc', 'def'], ['Client', 'Client'] )
	io_manager.read_file()

